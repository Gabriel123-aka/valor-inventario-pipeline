# =========================================================
# PROYECTO: Automatización Valor de Inventario
# VERSIÓN: 1.0.0 (Oficial)
# FECHA: 2025-12-24
# DESCRIPCIÓN: Sincronización completa de Portal Web y Excel.
#              Generación de gráficas y balances mensuales.
# =========================================================
import pandas as pd
import os
import shutil
from pathlib import Path
import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import actualizar_portal  # Así conectamos ambos archivos

# ==========================================================
# CONFIGURACIÓN PORTABLE Y SEGURIDAD (MODO DEMO)
# ==========================================================
MODO_DEMO = True  

UNC_FOLDER_ORIGEN = r"\\192.168.1.195\odbc_dir\Planeacion"
UNC_FOLDER_DESTINO = r"\\192.168.11.1\Planeacion\PublicaInventario"
BASE_DIR = Path(__file__).resolve().parent.parent

# --- Lógica de ORIGEN (Lectura de Excels) ---
# Forzamos el uso de data_samples siempre que MODO_DEMO sea True
UNC_FOLDER = str(BASE_DIR / "data_samples") if MODO_DEMO else UNC_FOLDER_ORIGEN
print(f"📂 Carpeta de lectura establecida en: {UNC_FOLDER}")

# --- Lógica de DESTINO (Seguridad del Portal) ---
if MODO_DEMO or not os.path.exists(UNC_FOLDER_DESTINO):
    CARPETA_DESTINO = str(BASE_DIR / "output")
    print(f"🚀 MODO DEMO ACTIVO: Resultados protegidos en carpeta local /output")
else:
    CARPETA_DESTINO = UNC_FOLDER_DESTINO
    print(f"🏢 MODO PRODUCCIÓN: Actualizando portal oficial en la red.")

os.makedirs(CARPETA_DESTINO, exist_ok=True)

# --- FUNCIÓN PARA COPIAR IMÁGENES AUTOMÁTICAMENTE ---
def asegurar_recursos_web(destino):
    ruta_web = os.path.join(BASE_DIR, "web")
    archivos = ["Tamex.jpg", "Almacen.png"] 
    for f in archivos:
        origen = os.path.join(ruta_web, f)
        meta = os.path.join(destino, f)
        if os.path.exists(origen):
            shutil.copy2(origen, meta)
            print(f"✅ Recurso copiado: {f}")

ARCHIVO_VALOR_INVENTARIO = os.path.join(CARPETA_DESTINO, "Valor de Inventario.xlsx")
HOJA_ANALISIS_GENERAL = "Analisis General"
HOJA_ABC = "ABC"
HOJA_TRANSITOS = "Transitos"
HOJA_DIAS_INV = "Dias Inventario"
HOJA_HISTORICO_CATEGORIA = "Historico Categoria"
HOJA_HISTORICO_ALMACEN = "Historico Almacen"
HOJA_COMPORTAMIENTO = "Comportamiento"
HOJA_RESUMEN_BALANCE = "Resumen y Balance"

PREFIXES = ["Inventario", "TransitosPendientes", "DOH_C", "OCPendiente", "Entradas X Planeacion"]
EXTS = [".xlsx", ".xls", ".csv"]

CLASIFICACIONES = ["NULL","A","B","C","D","E","I","N","X"]
OBJETIVO_CONSTANTE = 1875000000  

# =====================================================
# FUNCIONES AUXILIARES
# =====================================================
def fecha_hoy_str():
    return datetime.datetime.now().strftime("%Y%m%d")

def fecha_hoy_formato_ddmmyyyy():
    return datetime.datetime.now().strftime("%d/%m/%Y")

def encontrar_archivo(folder: str, prefix: str, date_str: str):
    folder_path = Path(folder)
    if not folder_path.exists():
        raise FileNotFoundError(f"La carpeta no existe o no es accesible: {folder}")
    pattern_base = f"{prefix} {date_str}"
    for ext in EXTS:
        matches = list(folder_path.glob(pattern_base + ext))
        if matches:
            return str(matches[0])
    matches_any = list(folder_path.glob(pattern_base + "*"))
    return str(matches_any[0]) if matches_any else None

def cargar_en_dataframe(path: str):
    if path is None:
        return None
    ext = Path(path).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(path)
    elif ext == ".csv":
        return pd.read_csv(path)
    try:
        return pd.read_excel(path)
    except:
        return pd.read_csv(path)

# =====================================================
# PROCESO PRINCIPAL
# =====================================================
def main():
    if MODO_DEMO:
        date_str = "20260206"  # Fecha fija para cuando no estemos en la red de la empresa, así mantenemos consistencia en los datos de ejemplo
        fecha_hoy = "06/02/2026"
        print(f"🚀 MODO DEMO: Tiempo congelado en {fecha_hoy} para consistencia de datos.")
    else:

        date_str = fecha_hoy_str()
        fecha_hoy = fecha_hoy_formato_ddmmyyyy()
        print(f"\n=== BUSCANDO ARCHIVOS DEL DÍA {date_str} ===\n")

    resultados = {}
    for pref in PREFIXES:
        ruta = encontrar_archivo(UNC_FOLDER, pref, date_str)
        if ruta is None:
            print(f"✖ No se encontró archivo: '{pref} {date_str}'\n")
            resultados[pref] = {"path": None, "df": None}
            continue
        print(f"✔ Encontrado: {ruta}")
        df = cargar_en_dataframe(ruta)
        resultados[pref] = {"path": ruta, "df": df}
        print(f"  → Cargado correctamente ({len(df)} filas, {len(df.columns)} columnas)\n")

    # =====================================================
    # 2. Agregar columna IMPORTE a Inventario
    # =====================================================
    df_inventario = resultados["Inventario"]["df"]
    if df_inventario is not None:
        columnas_necesarias = ["Existencias", "CostoPromedio", "TipoCambio"]
        faltantes = [c for c in columnas_necesarias if c not in df_inventario.columns]
        if faltantes:
            print(f"❌ No se puede crear 'Importe'. Faltan columnas: {faltantes}")
        else:
            for c in columnas_necesarias:
                df_inventario[c] = pd.to_numeric(df_inventario[c], errors="coerce").fillna(0)
            df_inventario["Importe"] = (
                df_inventario["Existencias"]
                * df_inventario["CostoPromedio"]
                * df_inventario["TipoCambio"]
            )
            df_inventario["Importe"] = df_inventario["Importe"].apply(lambda x: x if pd.notna(x) and x != 0 else "NULL")
            print("✔ Se agregó la columna 'Importe' y se reemplazaron vacíos por 'NULL'.")

    # =====================================================
    # 3. Previsualización
    # =====================================================
    for pref in PREFIXES:
        df = resultados[pref]["df"]
        print(f"\n=== PREVISUALIZACIÓN: {pref} ===")
        if df is not None:
            print(df.head(10))
        else:
            print(f"No se cargó ningún DataFrame para {pref}.")
    if df_inventario is not None and "Importe" in df_inventario.columns:
        suma_importe = df_inventario.loc[df_inventario["Importe"]!="NULL","Importe"].sum()
        print(f"\n💰 Suma total de 'Importe' en Inventario (ignorando NULL): {suma_importe:,.2f}")

    # =====================================================
    # 4. Actualizar Analisis General
    # =====================================================
    if df_inventario is not None and os.path.exists(ARCHIVO_VALOR_INVENTARIO):
        try:
            df_valor = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_ANALISIS_GENERAL)
            df_valor.columns = df_valor.columns.str.strip()
            columnas_valor = ["Tipo de Almacen", "Almacen", "IMPORTE"]
            faltantes_valor = [c for c in columnas_valor if c not in df_valor.columns]
            if faltantes_valor:
                print(f"❌ No se puede actualizar Analisis General. Faltan columnas: {faltantes_valor}")
            else:
                df_agg = df_inventario[df_inventario["Importe"]!="NULL"].groupby("Almacen")["Importe"].sum().reset_index()
                df_valor["IMPORTE"] = df_valor["Almacen"].map(df_agg.set_index("Almacen")["Importe"]).fillna(0)
                with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_valor.to_excel(writer, sheet_name=HOJA_ANALISIS_GENERAL, index=False)
                print(f"✔ Hoja '{HOJA_ANALISIS_GENERAL}' actualizada.")
        except Exception as e:
            print(f"❌ Error al actualizar hoja Analisis General: {e}")

    # =====================================================
    # 5. Actualizar ABC
    # =====================================================
    try:
        df_inventario["Almacen_norm"] = df_inventario["Almacen"].astype(str).str.strip().str.upper()
        df_inventario["ABCGeneral_norm"] = (
            df_inventario["ABCGeneral"]
            .astype(str)
            .str.strip()
            .str.upper()
            .replace({"": "NULL", "NAN": "NULL", "NONE": "NULL"})
        )
        df_inventario["Importe_n"] = pd.to_numeric(df_inventario["Importe"], errors="coerce").fillna(0)

        df_sumas = (
            df_inventario.groupby(["Almacen_norm", "ABCGeneral_norm"])["Importe_n"]
            .sum()
            .reset_index()
        )

        df_abc = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_ABC)
        df_abc["Almacen_norm"] = df_abc["Almacen"].astype(str).str.strip().str.upper()

        for idx, row in df_abc.iterrows():
            almacen = row["Almacen_norm"]
            if almacen == "TOTAL":
                continue
            for clas in CLASIFICACIONES:
                encontrado = df_sumas[
                    (df_sumas["Almacen_norm"] == almacen) &
                    (df_sumas["ABCGeneral_norm"] == clas)
                ]
                valor = encontrado["Importe_n"].values[0] if not encontrado.empty else 0
                df_abc.at[idx, clas] = valor

        total_idx = df_abc[df_abc["Almacen_norm"] == "TOTAL"].index[0]
        for clas in CLASIFICACIONES:
            df_abc.at[total_idx, clas] = df_abc.loc[df_abc["Almacen_norm"]!="TOTAL", clas].sum()
        df_abc["TOTAL"] = df_abc[CLASIFICACIONES].sum(axis=1)
        df_abc = df_abc.drop(columns=["Almacen_norm"], errors="ignore")

        with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                            mode="a", if_sheet_exists="replace") as writer:
            df_abc.to_excel(writer, sheet_name=HOJA_ABC, index=False)

        print("✔ Hoja ABC actualizada correctamente.")

    except Exception as e:
        print(f"❌ Error al actualizar hoja ABC: {e}")

    # =====================================================
    # 6. Actualizar hoja Transitos
    # =====================================================
    df_transitos = resultados["TransitosPendientes"]["df"]
    if df_transitos is not None:
        try:
            df_transitos = df_transitos[df_transitos["Mov"].str.strip().str.upper() == "TRANSITO"].copy()

            df_transitos["CantidadPendiente"] = pd.to_numeric(df_transitos["CantidadPendiente"], errors="coerce").fillna(0)
            df_transitos["Costo"] = pd.to_numeric(df_transitos["Costo"], errors="coerce").fillna(0)
            df_transitos["IMPORTE"] = df_transitos["CantidadPendiente"] * df_transitos["Costo"]
            df_transitos["Proyecto"] = ""

            columnas_finales = ["Mov","MovID","Estatus","FechaEmision","Articulo","Descripcion1",
                                "Cantidad","AlmacenPartida","AlmacenDestino","Costo","Observaciones",
                                "CantidadPendiente","Proyecto","IMPORTE"]

            df_transitos = df_transitos[[c for c in columnas_finales if c in df_transitos.columns]]

            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                                mode="a", if_sheet_exists="replace") as writer:
                df_transitos.to_excel(writer, sheet_name=HOJA_TRANSITOS, index=False)

            print("✔ Hoja Transitos actualizada correctamente.")

        except Exception as e:
            print(f"❌ Error al actualizar hoja Transitos: {e}")

    # =====================================================
    # 7. Actualizar hoja Dias Inventario
    # =====================================================
    df_doh = resultados["DOH_C"]["df"]
    totales = {} # Definir fuera para acceso global en el script
    if df_doh is not None:
        try:
            for col in ["Disponible", "Transitos", "Venta", "OCompra", "PedidosP"]:
                df_doh[col] = pd.to_numeric(df_doh[col], errors="coerce").fillna(0)

            df_doh["VPM"] = df_doh["Venta"] / 12
            df_doh["PROYECCION"] = df_doh["Disponible"] + df_doh["Transitos"] + df_doh["OCompra"] - df_doh["PedidosP"]

            df_doh["DOH_PROY"] = df_doh.apply(
                lambda r: "Sin Venta" if r["VPM"] == 0 else
                (r["Disponible"] + r["Transitos"] + r["OCompra"] - r["PedidosP"]) / (r["VPM"] * 12 / 360),
                axis=1
            )

            df_doh["DOH_INM"] = df_doh.apply(
                lambda r: "Sin Venta" if r["VPM"] == 0 else
                ((r["Disponible"] + r["Transitos"]) / r["VPM"]) * 30,
                axis=1
            )

            df_doh["MESES_FINAL"] = df_doh["DOH_PROY"].apply(lambda val: val if val=="Sin Venta" else val/30)

            df_doh = df_doh.sort_values(by="Disponible", ascending=False).reset_index(drop=True)

            for col in ["Disponible", "Transitos", "Venta", "OCompra", "PedidosP", "VPM", "PROYECCION"]:
                totales[col] = df_doh[col].sum()
            totales["DOH_PROY"] = (totales["Disponible"] + totales["Transitos"] + totales["OCompra"] - totales["PedidosP"]) / (totales["VPM"] * 12 / 360) if totales["VPM"] !=0 else "Sin Venta"
            totales["DOH_INM"] = ((totales["Disponible"] + totales["Transitos"]) / totales["VPM"]) * 30 if totales["VPM"] !=0 else "Sin Venta"
            totales["MESES_FINAL"] = totales["DOH_PROY"]/30 if totales["DOH_PROY"] != "Sin Venta" else "Sin Venta"

            df_doh.loc[len(df_doh)] = [ "TOTAL" ] + [totales[col] for col in ["Disponible", "Transitos", "Venta", "OCompra", "PedidosP", "VPM", "PROYECCION","DOH_PROY","DOH_INM","MESES_FINAL"]]

            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                                mode="a", if_sheet_exists="replace") as writer:
                df_doh.to_excel(writer, sheet_name=HOJA_DIAS_INV, index=False)

            print("✔ Hoja Dias Inventario actualizada correctamente.")

        except Exception as e:
            print(f"❌ Error al actualizar hoja Dias Inventario: {e}")

    # =====================================================
    # 8. Actualizar Historico Categoria
    # =====================================================
    if df_inventario is not None and os.path.exists(ARCHIVO_VALOR_INVENTARIO):
        try:
            df_analisis = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_ANALISIS_GENERAL)
            df_analisis["Tipo de Almacen"] = df_analisis["Tipo de Almacen"].astype(str).str.strip().str.upper()
            almacenes_validos = ["ALMACENES FACTURACIÓN", "ALMACENES CONSIGNACION", "ALMACENES MALESTADO"]
            almacenes_filtrados = df_analisis[df_analisis["Tipo de Almacen"].isin(almacenes_validos)]["Almacen"].tolist()

            df_inv_filtrado = df_inventario[df_inventario["Almacen"].isin(almacenes_filtrados)]
            df_inv_filtrado = df_inv_filtrado[df_inv_filtrado["Importe"] != "NULL"]

            df_hist = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_HISTORICO_CATEGORIA)

            df_sum = df_inv_filtrado.groupby("Categoria")["Importe"].sum().reset_index()

            col_fecha = fecha_hoy

            if col_fecha not in df_hist.columns:
                insert_pos = df_hist.columns.get_loc("Categoria") + 1
                df_hist.insert(insert_pos, col_fecha, 0)

            for idx, row in df_sum.iterrows():
                cat = row["Categoria"]
                valor = row["Importe"]
                if cat in df_hist["Categoria"].values:
                    df_hist.loc[df_hist["Categoria"] == cat, col_fecha] = valor
                else:
                    nueva_fila = {c: 0 for c in df_hist.columns}
                    nueva_fila["Categoria"] = cat
                    nueva_fila[col_fecha] = valor
                    df_hist = pd.concat([df_hist, pd.DataFrame([nueva_fila])], ignore_index=True)

            df_hist[col_fecha] = df_hist[col_fecha].fillna(0)
            df_hist = df_hist.sort_values(by=col_fecha, ascending=False).reset_index(drop=True)

            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                                mode="a", if_sheet_exists="replace") as writer:
                df_hist.to_excel(writer, sheet_name=HOJA_HISTORICO_CATEGORIA, index=False)

            print(f"✔ Hoja '{HOJA_HISTORICO_CATEGORIA}' actualizada correctamente y ordenada de mayor a menor por '{col_fecha}'.")

        except Exception as e:
            print(f"❌ Error al actualizar hoja Historico Categoria: {e}")

    # =====================================================
    # 9. Actualizar Historico Almacen
    # =====================================================
    if df_inventario is not None and os.path.exists(ARCHIVO_VALOR_INVENTARIO):
        try:
            df_hist_alm = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_HISTORICO_ALMACEN)

            col_fecha = fecha_hoy

            if col_fecha not in df_hist_alm.columns:
                insert_pos = df_hist_alm.columns.get_loc("Almacen") + 1
                df_hist_alm.insert(insert_pos, col_fecha, 0)

            for idx, row in df_hist_alm.iterrows():
                almacen = row["Almacen"]
                valor = df_inventario.loc[df_inventario["Almacen"] == almacen, "Importe"]
                valor = valor[valor != "NULL"].sum() if not valor.empty else 0
                df_hist_alm.loc[idx, col_fecha] = valor

            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                                mode="a", if_sheet_exists="replace") as writer:
                df_hist_alm.to_excel(writer, sheet_name=HOJA_HISTORICO_ALMACEN, index=False)

            print(f"✔ Hoja '{HOJA_HISTORICO_ALMACEN}' actualizada correctamente.")

        except Exception as e:
            print(f"❌ Error al actualizar hoja Historico Almacen: {e}")

    # =====================================================
    # 10. Actualizar hoja Comportamiento (fila más reciente arriba)
    # =====================================================
    try:
        df_comport = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_COMPORTAMIENTO)
        df_analisis = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_ANALISIS_GENERAL)

        almacenes_principales = [
            "ALMACENES FACTURACIÓN",
            "ALMACENES CONSIGNACION",
            "ALMACENES MALESTADO"
        ]

        # Solo sumar ABC + DOH + STOCK de esos 3 tipos
        df_analisis_filtrado = df_analisis[df_analisis["Tipo de Almacen"].str.strip().str.upper().isin(almacenes_principales)]

        valor_inventario_total = df_analisis_filtrado["IMPORTE"].sum()
        valor_transitos_total = df_transitos["IMPORTE"].sum() if df_transitos is not None else 0
        valor_total_dia = valor_inventario_total + valor_transitos_total

        doh_proy_total = totales.get("DOH_PROY", 0)
        objetivo = OBJETIVO_CONSTANTE

        # --- NUEVA LÓGICA: no duplicar fecha ---
        if (df_comport["Fecha"].astype(str).str.strip() == fecha_hoy).any():
            print(f"ℹ Ya existe un registro para la fecha {fecha_hoy}, no se agregará fila duplicada.")
            # NUEVA LÍNEA: Buscamos el valor que ya existe en el DataFrame para usarlo en el portal
            variacion_diaria = df_comport.loc[df_comport["Fecha"].astype(str).str.strip() == fecha_hoy, "Variacion Diaria"].values[0]


        else:
            ultimo_valor = df_comport.iloc[0]["Valor Total"] if not df_comport.empty else 0
            variacion_diaria = valor_total_dia - ultimo_valor

            fila_nueva = {
                "Fecha": fecha_hoy,
                "Valor Total": valor_total_dia,
                "DOH Proyectado": doh_proy_total,
                "Objetivo": objetivo,
                "Variacion Diaria": variacion_diaria
            }

            df_comport = pd.concat([pd.DataFrame([fila_nueva]), df_comport], ignore_index=True)

        df_comport["Fecha"] = df_comport["Fecha"].apply(
            lambda x: x.strftime("%d/%m/%Y") if isinstance(x, (datetime.datetime, pd.Timestamp)) else x
        )

        with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl",
                            mode="a", if_sheet_exists="replace") as writer:
            df_comport.to_excel(writer, sheet_name=HOJA_COMPORTAMIENTO, index=False)

        print(f"✔ Hoja '{HOJA_COMPORTAMIENTO}' actualizada correctamente con la fila del día en curso arriba.")

    except Exception as e:
        print(f"❌ Error al actualizar hoja Comportamiento: {e}")

    # =====================================================
    # 11. Generar gráfica mensual Valor Total vs Objetivo
    # =====================================================
    try:
        # =======================================================
        # VARIABLE DE CONTROL
        # =======================================================
        DIAS_A_MOSTRAR = 7 # Define el número de días a incluir en la gráfica (Día actual + 6 días anteriores)
        
        # Leer hoja Comportamiento
        df_comp = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_COMPORTAMIENTO)

        # Convertir fechas
        df_comp["Fecha"] = pd.to_datetime(df_comp["Fecha"], format="%d/%m/%Y", errors="coerce")
        
        # 1. Ordenar por fecha ascendente para la gráfica (el más antiguo primero)
        df_comp = df_comp.sort_values("Fecha", ascending=True)

        # APLICAR FILTRO DE DÍAS USANDO LA VARIABLE
        df_ultimos_dias = df_comp.tail(DIAS_A_MOSTRAR).copy()

        fecha_max = df_ultimos_dias["Fecha"].max()
        
        # Crear columna Etiqueta X (solo el número del día)
        df_ultimos_dias["EtiquetaX"] = df_ultimos_dias["Fecha"].dt.strftime("%d-%b") 
        
        # 3. Asignar datos diarios
        x_labels = df_ultimos_dias["EtiquetaX"]
        valores = df_ultimos_dias["Valor Total"]
        objetivo_plot = df_ultimos_dias["Objetivo"] 
        doh_proyectado = pd.to_numeric(df_ultimos_dias["DOH Proyectado"], errors='coerce').fillna(0)

        # 4. Crear la gráfica y los DOS EJES Y
        fig, ax1 = plt.subplots(figsize=(12, 6)) 
        
        # ax1 será el eje primario (Valor Total / Objetivo)
        ax1.plot(x_labels, valores, marker="o", label="Valor Total", color='C0', linewidth=1.0, alpha=0.6)
        ax1.plot(x_labels, objetivo_plot, marker="x", linestyle="--", label="Objetivo", color='C1', linewidth=1.0, alpha=0.6)

        # ax2 será el eje secundario (DOH Proyectado)
        ax2 = ax1.twinx() 
        ax2.plot(x_labels, doh_proyectado, marker="^", linestyle="-.", label="DOH Proyectado", color='C2', linewidth=1.0, alpha=0.6)
        
        # --- Configuración Ejes y Etiquetas ---
        
        # Configurar Eje X (Común)
        ax1.set_xlabel("Día")
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, linestyle=':', alpha=0.5, color='gray')
        
        # Configurar Eje Y Primario (Valor Total / Objetivo)
        ax1.set_ylabel("Valor Total")
        
        # Definir la función de formato para MILLONES ($M)
        def format_millions(x, pos):
            if x >= 1e6:
                return f'${x*1e-6:,.0f}M' 
            elif x >= 1e3:
                return f'${x*1e-3:1.0f}K'
            else:
                return f'${x:1.0f}'

        formatter = ticker.FuncFormatter(format_millions)
        ax1.yaxis.set_major_formatter(formatter)
        
        # Configurar Eje Y Secundario (DOH Proyectado)
        ax2.set_ylabel("DOH Proyectado (Días)", color='C2')
        ax2.tick_params(axis='y', labelcolor='C2')
        ax2.grid(False) 

        # 5. Título y Leyenda Unificada
        plt.title(f"Valor Total, Objetivo y DOH - Comportamiento Diario (Últimos {DIAS_A_MOSTRAR} días)")
        
        # Mover leyenda fuera del gráfico
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, 
                loc='upper left', 
                bbox_to_anchor=(1.05, 1.05)) 
        
        # --- Etiquetas de Datos (Destacadas) ---
        
        def label_format(val):
            return f'{val:,.0f}'
            
        # Etiquetas para Valor Total (Eje 1)
        for i, val in enumerate(valores):
            ax1.text(i, val * 1.0005, label_format(val), ha='center', va='bottom', fontsize=7, color='C0', fontweight='bold')
            
        # Etiquetas para Objetivo (Eje 1)
        for i, obj in enumerate(objetivo_plot):
            ax1.text(i, obj * 0.995, label_format(obj), ha='center', va='top', fontsize=8, color='C1', fontweight='bold')
            
        # Etiquetas para DOH Proyectado
        for i, doh in enumerate(doh_proyectado):
            ax2.text(i, doh * 1.0005, f'{doh:,.2f}', ha='center', va='bottom', fontsize=8, color='C2', fontweight='bold') 

        # Ajustar límite Y primario
        y_min_ax1 = min(valores.min(), objetivo_plot.min())
        y_max_ax1 = max(valores.max(), objetivo_plot.max())
        ax1.set_ylim(y_min_ax1 * 0.95, y_max_ax1 * 1.05) 

        # --- Guardado e Inserción ---
        
        # 1. Ruta absoluta a tu carpeta de prueba
        ruta_grafica_final = os.path.join(CARPETA_DESTINO, "grafica_comportamiento.png")
        
        # 2. Guardar la imagen físicamente
        fig.canvas.draw()
        plt.tight_layout()
        fig.savefig(ruta_grafica_final, dpi=150, bbox_inches='tight', facecolor='white')

        plt.close(fig)

        print(f"✔ Gráfica guardada físicamente en: {ruta_grafica_final}")

       
    except Exception as e:
        print(f"❌ Error al generar la imagen de la gráfica: {e}")

    # =====================================================
    # 12. Actualizar hoja Resumen y Balance con Métricas Clave
    # =====================================================
    if os.path.exists(ARCHIVO_VALOR_INVENTARIO):
        try:
            # Se unifica el proceso de escritura de métricas en este bloque
            wb = load_workbook(ARCHIVO_VALOR_INVENTARIO)
            
            if HOJA_RESUMEN_BALANCE not in wb.sheetnames:
                print(f"❌ Error: La hoja '{HOJA_RESUMEN_BALANCE}' no existe en el archivo.")
                return 
                
            ws = wb[HOJA_RESUMEN_BALANCE]

            # Se agregara le fecha en las Metricas Clave del Día
            ws["B3"] = fecha_hoy
            ws["B3"].number_format = "DD/MM/YYYY"

            # Diccionario de valores a escribir (Celda: Valor)
            valores_clave = {
                "B5": valor_transitos_total,       # Valor Transito
                "B6": valor_inventario_total,      # Valor Inventario Físico
                "B7": valor_total_dia,             # Valor Total
                "B8": OBJETIVO_CONSTANTE,          # Objetivo
                "B9": totales.get("DOH_PROY", "N/A"), # DOH Proyectado
                "B10": totales.get("DOH_INM", "N/A") # DOH Inmediato
            }
            
            # Formato de moneda para los valores
            currency_format = '"$"#,##0.00'
            
            for celda, valor in valores_clave.items():
                if isinstance(valor, (float, int)):
                    ws[celda] = valor
                    if celda in ["B5", "B6", "B7", "B8"]:
                         ws[celda].number_format = currency_format
                    elif celda in ["B9", "B10"]:
                         ws[celda].number_format = "0.00"
                else:
                    ws[celda] = str(valor)
            wb.save(ARCHIVO_VALOR_INVENTARIO)
            print(f"✔ Hoja '{HOJA_RESUMEN_BALANCE}' actualizada con métricas clave.")

        except Exception as e:
            print(f"❌ Error al actualizar hoja Resumen y Balance (Métricas): {e}")

    # =====================================================
    # 13. Actualizar Balance Mensual (Últimos 3 meses) e Insertar Gráfica
    # =====================================================
    if os.path.exists(ARCHIVO_VALOR_INVENTARIO):
        try:
            # 1. CÁLCULO: Leer datos de Comportamiento y agrupar por Mes-Año
            df_comp_mensual = pd.read_excel(ARCHIVO_VALOR_INVENTARIO, sheet_name=HOJA_COMPORTAMIENTO)
            
            # Asegurar que la columna Fecha sea datetime y Variacion Diaria sea numérica
            df_comp_mensual["Fecha"] = pd.to_datetime(df_comp_mensual["Fecha"], format="%d/%m/%Y", errors="coerce")
            df_comp_mensual["Variacion Diaria"] = pd.to_numeric(df_comp_mensual["Variacion Diaria"], errors="coerce").fillna(0)

            df_comp_mensual["MesAnio"] = df_comp_mensual["Fecha"].dt.strftime("%b-%y")
            df_mensual_agg = df_comp_mensual.groupby("MesAnio")["Variacion Diaria"].sum().reset_index()

            # 2. CÁLCULO: Generar el rango dinámico de 3 meses (MES ACTUAL + 2 ANTERIORES)
            fecha_max_data = df_comp_mensual["Fecha"].max() 
            rango_meses_period = pd.period_range(end=fecha_max_data.to_period('M'), periods=3, freq='M')
            rango_meses_str = [p.strftime("%b-%y") for p in rango_meses_period]
            
            df_resultado_mensual = pd.DataFrame(rango_meses_str, columns=["MesAnio"])
            df_resultado_mensual = pd.merge(
                df_resultado_mensual, 
                df_mensual_agg, 
                on="MesAnio", 
                how="left"
            ).fillna(0)
            
            # 3. ESCRITURA EN EXCEL: Datos y Gráfica
            wb = load_workbook(ARCHIVO_VALOR_INVENTARIO)
            ws = wb[HOJA_RESUMEN_BALANCE]

            # LIMPIEZA: Borrar rango antiguo (A14:B25) para asegurar que no queden datos de meses viejos
            for row_clean in range(14, 26):
                ws[f'A{row_clean}'] = None
                ws[f'B{row_clean}'] = None
            
            # Escritura de Meses y Valores (A14:B16 - últimos 3 meses)
            currency_format = '"$"#,##0.00' 
            for i, row_data in df_resultado_mensual.iterrows():
                fila_actual = 14 + i
                ws.cell(row=fila_actual, column=1).value = row_data["MesAnio"]
                ws.cell(row=fila_actual, column=2).value = row_data["Variacion Diaria"]
                ws.cell(row=fila_actual, column=2).number_format = currency_format
            # --- INSERTAR LOGO TAMEX EN A1 ---
            ruta_logo = os.path.join(BASE_DIR, "web", "Tamex.jpg")
            if os.path.exists(ruta_logo):
                img_logo = XLImage(ruta_logo)
                # Escala pequeña para que quepa en la celda A1 (ajusta ancho/alto si es necesario)
                img_logo.width = 110 
                img_logo.height = 55
                ws.add_image(img_logo, "A1")
                print("✔ Logo Tamex insertado en A1.")    

            # INSERTAR GRÁFICA EXCLUSIVAMENTE AQUÍ
            if ws._images:
                ws._images.clear() 
                # Re-agregamos el logo después de limpiar
                if os.path.exists(ruta_logo):
                    img_logo = XLImage(ruta_logo)
                    img_logo.width = 140; img_logo.height = 85
                    ws.add_image(img_logo, "A1")
            

            # --- CORRECCIÓN UNIFICADA ---
            ruta_grafica_final = os.path.join(CARPETA_DESTINO, "grafica_comportamiento.png")
            if os.path.exists(ruta_grafica_final):
                img_comp = XLImage(ruta_grafica_final)
                img_comp.width = 700 
                img_comp.height = 350
                ws.add_image(img_comp, "D3") 
                print(f"✔ Gráfica sincronizada insertada en Excel desde: {ruta_grafica_final}")

            wb.save(ARCHIVO_VALOR_INVENTARIO)
        except Exception as e:
            print(f"❌ Error al actualizar Balance Mensual y Gráfica: {e}")

    # =====================================================
    # 14. Actualizar OCPendientes por Proveedor (Tabla Dinámica Mejorada)
    # =====================================================
    df_oc = resultados["OCPendiente"]["df"]
    if df_oc is not None:
        try:
            print("--- Procesando OCPendientes por Proveedor ---")
            
            # 1. Asegurar tipos de datos numéricos
            df_oc["ImportePendiente"] = pd.to_numeric(df_oc["ImportePendiente"], errors="coerce").fillna(0)
            df_oc["TipoCambio"] = pd.to_numeric(df_oc["TipoCambio"], errors="coerce").fillna(0)
            
            # 2. Crear columna calculada
            df_oc["Importe pendiente OK"] = df_oc["ImportePendiente"] * df_oc["TipoCambio"]
            
            # 3. Limpiar columna Proyecto y pasar a MINÚSCULAS
            df_oc["Proyecto"] = df_oc["Proyecto"].fillna("sin asignar").replace("", "sin asignar").astype(str).str.strip().str.lower()
            
            # 4. Crear la Pivot Table
            resumen_oc = df_oc.pivot_table(
                index="Nombre Proveedor", 
                columns="Proyecto", 
                values="Importe pendiente OK", 
                aggfunc="sum", 
                fill_value=0
            )
            
            # 5. Calcular la columna de Total Horizontal (IMPORTE PENDIENTE)
            # Al no hacer reset_index aún, Nombre Proveedor es el índice, no una columna.
            resumen_oc["IMPORTE PENDIENTE"] = resumen_oc.sum(axis=1)
            
            # 6. Ordenar de mayor a menor por el total
            resumen_oc = resumen_oc.sort_values(by="IMPORTE PENDIENTE", ascending=False)
            
            # 7. Ahora sí, bajamos el índice a columna y renombramos
            resumen_oc = resumen_oc.reset_index()
            resumen_oc = resumen_oc.rename(columns={"Nombre Proveedor": "NOMBRE DE PROVEEDOR"})
            
            # Convertimos los nombres de proveedores a MAYÚSCULAS
            resumen_oc["NOMBRE DE PROVEEDOR"] = resumen_oc["NOMBRE DE PROVEEDOR"].astype(str).str.upper()

            # --- REORDENAR: Solo las columnas necesarias (Elimina cualquier duplicada) ---
            # Identificamos los nombres de los proyectos (en minúsculas)
            cols_proyectos = [c for c in resumen_oc.columns if c not in ["NOMBRE DE PROVEEDOR", "IMPORTE PENDIENTE"]]
            
            # Forzamos el orden: PROVEEDOR (A), TOTAL (B), PROYECTOS (C en adelante)
            columnas_finales = ["NOMBRE DE PROVEEDOR", "IMPORTE PENDIENTE"] + cols_proyectos
            resumen_oc = resumen_oc[columnas_finales]

            # 8. Agregar Fila de TOTALES al final
            fila_totales = {"NOMBRE DE PROVEEDOR": "TOTAL GENERAL"}
            for col_tot in resumen_oc.columns:
                if col_tot != "NOMBRE DE PROVEEDOR":
                    fila_totales[col_tot] = resumen_oc[col_tot].sum()
            
            resumen_oc = pd.concat([resumen_oc, pd.DataFrame([fila_totales])], ignore_index=True)

            # 9. Guardar en Excel y aplicar FORMATO
            HOJA_OC_PROV_NAME = "OCPendientes por Proveedor"
            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                resumen_oc.to_excel(writer, sheet_name=HOJA_OC_PROV_NAME, index=False)
                
                ws_oc = writer.sheets[HOJA_OC_PROV_NAME]
                from openpyxl.styles import Font
                
                for col_cells in ws_oc.columns:
                    max_len = 0
                    column_letter = col_cells[0].column_letter
                    header_value = col_cells[0].value 
                    
                    for cell in col_cells:
                        # 1. Encabezados: Negrita solo para Proveedor e Importe Pendiente
                        if cell.row == 1:
                            if header_value in ["NOMBRE DE PROVEEDOR", "IMPORTE PENDIENTE"]:
                                cell.font = Font(bold=True)
                            else:
                                cell.font = Font(bold=False) # Proyectos en minúsculas sin negrita

                        # 2. Formato de moneda para valores numéricos
                        if cell.row > 1 and isinstance(cell.value, (int, float)):
                            cell.number_format = '"$"#,##0.00'
                        
                        # 3. Fila de TOTAL GENERAL (siempre en negrita)
                        if ws_oc.cell(row=cell.row, column=1).value == "TOTAL GENERAL":
                            cell.font = Font(bold=True)

                        # 4. Cálculo de ancho para auto-ajuste
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except: pass
                    
                    ws_oc.column_dimensions[column_letter].width = max_len + 4

            print(f"✔ Hoja '{HOJA_OC_PROV_NAME}' actualizada. Columna duplicada eliminada.")

        except Exception as e:
            print(f"❌ Error al procesar OCPendientes por Proveedor: {e}")
    
    # =====================================================
    # 15. Actualizar hoja Entradas X Planeación (Copia Espejo)
    # =====================================================
    df_entradas = resultados["Entradas X Planeacion"]["df"]
    if df_entradas is not None:
        try:
            print("--- Procesando Entradas X Planeación ---")
            HOJA_ENTRADAS_NAME = "Entradas X Planeación"
            
            with pd.ExcelWriter(ARCHIVO_VALOR_INVENTARIO, engine="openpyxl", 
                                mode="a", if_sheet_exists="replace") as writer:
                df_entradas.to_excel(writer, sheet_name=HOJA_ENTRADAS_NAME, index=False)
                
                # Auto-ajuste de columnas
                ws_ent = writer.sheets[HOJA_ENTRADAS_NAME]
                for col_cells in ws_ent.columns:
                    max_len = 0
                    column_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            if cell.value:
                                length = len(str(cell.value))
                                if length > max_len: max_len = length
                        except: pass
                    ws_ent.column_dimensions[column_letter].width = max_len + 2

            print(f"✔ Hoja '{HOJA_ENTRADAS_NAME}' actualizada correctamente.")

        except Exception as e:
            print(f"❌ Error al procesar la hoja Entradas X Planeación: {e}")
    
    # =====================================================
    # 15.1. TOP 10 ENTRADAS (NUEVA SECCIÓN)
    # =====================================================
    if df_entradas is not None:
        try:
            print("--- Generando Top 10 de Entradas ---")
            
            # 1. Asegurar que Importe2 sea numérico
            df_entradas["Importe2"] = pd.to_numeric(df_entradas["Importe2"], errors="coerce").fillna(0)

            # Buscar la última fecha con datos que no sea hoy
            df_entradas['FechaEmision'] = pd.to_datetime(df_entradas['FechaEmision'], errors='coerce')
            fecha_ayer = df_entradas[df_entradas['FechaEmision'].dt.date < pd.to_datetime(date_str,format='%Y%m%d').date()]['FechaEmision'].max()
            
            # Crear el recorte del último día operativo
            df_solo_ayer = df_entradas[df_entradas['FechaEmision'] == fecha_ayer]
            
            # 2. Agrupar, Sumar y Ordenar
            top_10_df = df_solo_ayer.groupby("Nombre")["Importe2"].sum().reset_index()
            top_10_df = top_10_df.sort_values(by="Importe2", ascending=False).head(10)

            # 3. Cargar el Workbook para escribir en celdas específicas
            wb = load_workbook(ARCHIVO_VALOR_INVENTARIO)
            ws_resumen = wb[HOJA_RESUMEN_BALANCE]

            # 4. Limpiar rango antiguo (A21:B30)
            for r in range(21, 31):
                ws_resumen[f"A{r}"] = None
                ws_resumen[f"B{r}"] = None

            # 5. Escribir los datos y aplicar formato
            currency_format = '"$"#,##0.00'
            for i, (idx, row) in enumerate(top_10_df.iterrows()):
                fila = 21 + i
                ws_resumen[f"A{fila}"] = str(row["Nombre"]).upper()
                ws_resumen[f"B{fila}"] = row["Importe2"]
                ws_resumen[f"B{fila}"].number_format = currency_format

            # 6. Ajustar ancho de columnas A y B
            if not top_10_df.empty:
                max_len_nombre = top_10_df["Nombre"].astype(str).map(len).max()
                ws_resumen.column_dimensions['A'].width = max_len_nombre + 5
            ws_resumen.column_dimensions['B'].width = 18

            # =====================================================
            # 15.2. CÁLCULO DE IMPORTES ADICIONALES (B32 y B34)
            # =====================================================
            print("--- Calculando Importes Ayer y Mes ---")
            
            # Asegurar formato de fecha
            df_entradas['FechaEmision'] = pd.to_datetime(df_entradas['FechaEmision'], errors='coerce')
            fecha_hoy_dt = pd.to_datetime(date_str, format='%Y%m%d')
            
            # A. Cálculo Día Anterior (B32)
            fechas_anteriores = df_entradas[df_entradas['FechaEmision'].dt.date < fecha_hoy_dt.date()]
            importe_dia_anterior = 0
            if not fechas_anteriores.empty:
                ultima_fecha = fechas_anteriores['FechaEmision'].max()
                importe_dia_anterior = fechas_anteriores[fechas_anteriores['FechaEmision'] == ultima_fecha]['Importe2'].sum()
                print(f"ℹ️ Día anterior detectado: {ultima_fecha.date()}")

            # B. Cálculo Acumulado Mes (B34)
            inicio_mes = fecha_hoy_dt.replace(day=1)
            importe_acumulado_mes = df_entradas[
                (df_entradas['FechaEmision'].dt.date >= inicio_mes.date()) & 
                (df_entradas['FechaEmision'].dt.date <= fecha_hoy_dt.date())
            ]['Importe2'].sum()

            # C. Escribir en Excel con formato
            ws_resumen["B32"] = importe_dia_anterior
            ws_resumen["B34"] = importe_acumulado_mes
            ws_resumen["B32"].number_format = currency_format
            ws_resumen["B34"].number_format = currency_format

            # 7. GUARDADO FINAL ÚNICO
            wb.save(ARCHIVO_VALOR_INVENTARIO)
            print("✔ Top 10 e Importes (B32, B34) actualizados correctamente.")

        except Exception as e:
            print(f"❌ Error al procesar sección de Entradas: {e}")
    # =====================================================
    # 16. ACTUALIZACIÓN DEL PORTAL WEB (NUEVA SECCIÓN)
    # =====================================================
    try:

        info_para_web = {
            'fecha': fecha_hoy,
            'v_total': valor_total_dia,         
            'v_transito': valor_transitos_total, 
            'v_fisico': valor_inventario_total, 
            'doh': totales.get("DOH_PROY", 0),
            'v_diaria': variacion_diaria,
            'e_ayer': importe_dia_anterior,
            'e_mes': importe_acumulado_mes,
            'top_10': top_10_df.to_dict(orient='records'),
            'm1_n': df_resultado_mensual.iloc[0]['MesAnio'], 
            'm1_v': df_resultado_mensual.iloc[0]['Variacion Diaria'],
            'm2_n': df_resultado_mensual.iloc[1]['MesAnio'], 
            'm2_v': df_resultado_mensual.iloc[1]['Variacion Diaria'],
            'm3_n': df_resultado_mensual.iloc[2]['MesAnio'], 
            'm3_v': df_resultado_mensual.iloc[2]['Variacion Diaria'],
            'ruta_destino': CARPETA_DESTINO      
        }

        # 2. Generar el index.html
        actualizar_portal.actualizar_index(info_para_web)
        # 1. Copiamos los recursos visuales a la carpeta final
        asegurar_recursos_web(CARPETA_DESTINO)

        print(f"✨ ¡Prueba generada! Revisa tu carpeta: {CARPETA_DESTINO}")
        
        print(f"✔ ¡Prueba generada! Revisa tu carpeta: {CARPETA_DESTINO}")
        
    except Exception as e:
        print(f"❌ Error en la actualización final: {e}")

# =====================================================
# EJECUCIÓN
# =====================================================
if __name__ == "__main__":
    main()
